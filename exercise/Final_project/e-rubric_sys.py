#!/usr/bin/python3
'''
===========================================
Jerry's e-Rubric system V0.9 alpha
Start time: 2018.6.12
Finish time of 1st generation: 2018.6.20
Enviroment required: Python 3.5.4rc1 with Windows or UNIX-like OS
Extension library: easygui & openpyxl
===========================================
'''

import os
import sys

import easygui as ui
import openpyxl as xl

# Test of os extension
print("sys:", os.name)


# find the abs path
abs_path = os.getcwd()
print(abs_path)  # Debugging purpose


# creat the folder
if os.path.exists(abs_path+"/rubric") == True:
    print("cwd:", True)  # Debugging purpose
else:
    os.mkdir(abs_path+"/rubric")

# change the working dir
os.chdir(abs_path+"/rubric")
abs_path = os.getcwd()


# Initialization the global variable
global name_lst
global final_mark_dict
global new_name
name_lst = []
final_mark_dict = {}
new_name = ""


def start_sence():
    global start_choice
    start_func = ["New Rubric", "Review Rubric", "About"]
    start_sentences = "Welcom to e-Rubric System, choose a funtion to start! (^.^)\n\nCurrent running directory is:" + \
        abs_path + "\n\nThis system is no longer support the old .xls files!" + \
        "\n\nPress About for help!"
    start_choice = ui.buttonbox(msg="{:^}".format(start_sentences),
                                title="Main Menu", choices=start_func)
    if start_choice == None:
        sys.exit()  # Exit the Program
    else:
        return start_choice  # return the choice


def new_sence():
    global name_lst
    message = "Do you want to creat a new rubric or continue editing a previous one?"
    choice = ["creat new rubric", "continue editing"]
    new_choice = ui.buttonbox(msg=message, choices=choice, title="New Rubric")
    # Calling the function
    if new_choice == "creat new rubric":
        # unavaliable of encryptiion
        passwd = ui.passwordbox(
            msg="Please set a password (unavaliable)", title="Protection")
        print(passwd)  # Debugging purpose
        new_rubric()  # Calling the method for creat empty file
        level = ui.enterbox(
            msg="Please input the total marks", title="New Rubric")
        name_sheet = ui.fileopenbox(msg="Please choose a students sheet", filetypes=[
                                    "*.xlsx"], default=r"\*.xlsx")
        wb = xl.load_workbook(filename=name_sheet)  # load Workbook
        a_sheet = wb.worksheets[0]  # load the first sheet
        # creat the name list
        for row in a_sheet.iter_rows():
            for cell in row:
                name_lst.append(cell.value)
        print(name_lst)  # Debugging purpose
        for name in name_lst:
            mark(name, level)  # Calling the marking method
        wirte_file(final_mark_dict, new_name)
    elif new_choice == "continue editing":
        # unavaliable
        new_sence()


def new_rubric():
    global new_name
    # Asking for the rubric's name
    new_name = ui.enterbox(msg="Naming the new rubric", title="New Rubric")
    exists_flag = os.path.exists(abs_path+"/"+new_name+".xlsx")
    # Dectect the file is or not exists
    if exists_flag == True:
        ui.msgbox(msg="Existed name, please choose another name!", title="ERROR")
        new_rubric()
    else:
        # creat empty file
        filepath = new_name+".xlsx"
        wb = xl.Workbook()
        wb.save(filepath)


def mark(name, total_level):
    global final_mark_dict
    level = []  # totall level of marking
    # creat marking sheet
    for i in range(int(total_level)+1):
        level.append(str(i))
    result = ui.choicebox(msg="Marking for student: " + name,
                          choices=level, title="Marking Activity")
    # storage result
    final_mark_dict[name] = result


def wirte_file(input_dict, file_name):
    global final_mark_dict
    # creat data dictionary
    names = list(final_mark_dict.keys())
    marks = list(final_mark_dict.values())
    # Initialization pointer
    row = 1
    wb = xl.load_workbook(filename=file_name+".xlsx")
    a_sheet = wb.worksheets[0]
    # wirte the names to the file
    for pointer in range(len(names)):
        a_sheet.cell(row=row, column=1).value = names[pointer]
        row += 1
    row = 1
    # wirte marks to the file
    for pointer in range(len(marks)):
        a_sheet.cell(row=row, column=2).value = marks[pointer]
        row += 1
    # save the file
    wb.save(file_name+".xlsx")
    ui.msgbox(msg="You are finished of marking the rubric, you could review it from the start sence!",
              title="New Rubric")
    choice = ui.ccbox(msg="Do you want to review the rubric now?",
                      title="continue", choices=("Y[e]s!", "N[o]!"))
    if choice == True:
        review()
    else:
        pass


def review():
    # Initialization the variable
    content = ""
    # choose the file to open (.xlsx)
    open_file = ui.fileopenbox(msg="Choose a rubric to review", title="Open", filetypes=[
                               "*.xlsx"], default=r"\*.xlsx")
    print("choosen:", open_file)  # Debugging purpose
    # open the file
    wb = xl.load_workbook(filename=open_file)
    a_sheet = wb.worksheets[0]
    # read & storage file to local variable
    for row in a_sheet.iter_rows():
        content = content + "\n"
        for cell in row:
            content = content + "   " + str(cell.value)
    # show the result
    ui.codebox(text=content, msg="Review the Previous Rubric\n\nFile directory is: " +
               open_file, title="Review the Rubric")


def about():
    ui.msgbox(
        msg="Powered by Jerry Xie in Maple Leaf International School.\n\nProgramming 11 Final Project\n\n2018.6.12", title="About")
    ui.msgbox(
        msg="Helping Manual of users:\n1. Please put every student's name in the first row. \n2. The mark of each students will be placed after his name.\n3. If the first row is not the name and the review function will crushed.\n4. The data that not in the first sheet will not be reconized.",
        title="Help")
    ui.abouteasygui()



# Main Window Loop
while True:
    start_sence()
    if start_choice == "New Rubric":
        new_sence()
    elif start_choice == "Review Rubric":
        review()
    elif start_choice == "About":
        about()
    else:
        sys.exit()


'''
zoom it smaller
                      :;J7, :,                        ::;7:
                      ,ivYi, ,                       ;LLLFS:
                      :iv7Yi                       :7ri;j5PL
                     ,:ivYLvr                    ,ivrrirrY2X,
                     :;r@checkJ:                :ivu@ingramD.
                    :iL7::,:::iiirii:ii;::::,,irvF7rvvLujL7ur
                   ri::,:,::i:iiiiiii:i:irrv177JX7rYXqZEkvv17
                ;i:, , ::::iirrririi:i:::iiir@ZOEii;L8OGJr71i
              :,, ,,:   ,::ir@johnX.irii:i:::j1jri7ZBOS7ivv,
                 ,::,    ::rv77iiiriii:iii:i::,rvLq@nidghoog
             ,,      ,, ,:ir7ir::,:::i;ir:::i:i::rSGGYri712:
           :::  ,v7r:: ::rrv77:, ,, ,:i7rrii:::::, ir7ri7Lri
          ,     2OBBOi,iiir;r::        ,irriiii::,, ,iv7Luur:
        ,,     i78MBBi,:,:::,:,  :7FSL: ,iriii:::i::,,:rLqXv::
        :      iuMMP: :,:::,:ii;2GY7OBB0viiii:i:iii:i:::iJqL;::
       ,     ::::i   ,,,,, ::LuBBu BBBBBErii:i:i:i:i:i:i:r77ii
      ,       :       , ,,:::rruBZ1MBBqi, :,,,:::,::::::iiriri:
     ,               ,,,,::::i:  @jerry.x       ,:,, ,:::ii;i7:
    :,       rjujLYLi   ,,:::::,:::::::::,,   ,:i,:,,,,,::i:iii
    ::      BBBBBBBBB0,    ,,::: , ,:::::: ,      ,,,, ,,:::::::
    i,  ,  ,8BMMBBBBBBi     ,,:,,     ,,, , ,   , , , :,::ii::i::
    :      iZMOMOMBBM2::::::::::,,,,     ,,,,,,:,,,::::i:irr:i:::,
    i   ,,:;u0MBMOG1L:::i::::::  ,,,::,   ,,, ::::::i:i:iirii:i:i:
    :    ,iuUuuXUkFu7i:iii:i:::, :,:,: ::::::::i:i:::::iirr7iiri::
    :     :rkY@mr.lee:::::, ,:ii:::::::i:::::i::,::::iirrriiiri::,
     :      5BMBBBBBBSr:,::rv2kuii:::iii::,:i:,, , ,,:,:i@marcoY.,
          , :r50EZ8MBBBBGOBBBZP7::::i::,:::::,: :,:,::i;rrririiii::
              :jujYY7LS0ujJL7r::,::i::,::::::::::::::iirirrrrrrr:ii:
           ,:  :@charlieC.:,:,,,::::i:i:::::,,::::::iir;ii;7v77;ii;i,
           ,,,     ,,:,::::::i:iiiii:i::::,, ::::iiiir@alcatraz.r;7:i,
        , , ,,,:,,::::::::iiiiiiiiii:,:,:::::::::iiir;ri7vL77rrirri::
         :,, , ::::::::i:::i:::i:i::,,,,,:,::i:i:::iir;@aslanW.ii:::
'''
